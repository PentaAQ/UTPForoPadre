from django.shortcuts import render
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from .models import Publicaciones, Categorias
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from .forms import ComentarioForm, PublicacionForm, CategoriaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models.functions import ExtractMonth
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import random



class PublicacionesListView(LoginRequiredMixin, ListView):
    template_name = "home.html"
    model = Publicaciones
    context_object_name = "publicaciones"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["mensaje"] = "Publicaciones disponibles"

        raw_counts = (
            self.model.objects.values("categoria__nombre")
            .annotate(total=Count("categoria"))
            .order_by("-total")[:3]
        )
        top_categories = [
            {"label": item["categoria__nombre"], "total": item["total"]}
            for item in raw_counts
        ]
        context["top_categories"] = top_categories
        return context


class ProductDetailView(DetailView):
    template_name = "publicacion.html"
    model = Publicaciones
    context_object_name = "publicacion"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ComentarioForm()
        context["comentarios"] = self.object.comentarios.all()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = ComentarioForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.publicacion = self.object
            comentario.autor = request.user
            comentario.save()
            messages.success(request, "Comentario agregado exitosamente")
            return redirect("detalle_publicacion", pk=self.object.pk)
        else:
            messages.error(request, "Error al agregar el comentario.")
            return self.render_to_response(self.get_context_data(form=form))

@login_required
def nueva_publicacion(request):
    form = PublicacionForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            publicacion = form.save(commit=False)
            publicacion.autor = request.user
            publicacion.save()
            messages.success(request, "Publicación creada exitosamente")
            return redirect("home")
        else:
            messages.error(
                request,
                "Error al crear la publicación. Por favor, revisa los datos ingresados.",
            )
    return render(request, "nueva_publicacion.html", {"form": form})


class MisPublicacionesListView(LoginRequiredMixin, ListView):
    template_name = "mispublicaciones.html"
    model = Publicaciones
    context_object_name = "publicaciones"

    def get_queryset(self):
        return Publicaciones.objects.filter(autor=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["mensaje"] = "Mis publicaciones"
        return context

@login_required
def lista_categorias(request):
    categorias = Categorias.objects.all()
    datos = []
    inicio_semana = timezone.now() - timedelta(days=timezone.now().weekday())
    for categoria in categorias:
        total = categoria.publicaciones_set.count()
        esta_semana = categoria.publicaciones_set.filter(fecha_creacion__gte=inicio_semana).count()
        if total >= 10:
            estrellas = 5
        else:
            estrellas = max(1, total // 2)  
        datos.append({
            'cat': categoria,
            'total': total,
            'esta_semana': esta_semana,
            'estrellas': estrellas,
        })
    totales = sum(item['total'] for item in datos)
    num_categorias = len(datos)
    promedio = totales / num_categorias if num_categorias > 0 else 0
    return render(request, "listacategorias.html", {"categorias": datos, "promedio": promedio})

@login_required
def nueva_categoria(request):
    form = CategoriaForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.save()
            messages.success(request, "Categoria creada exitosamente")
            return redirect("lista_categorias")
        else:
            messages.error(
                request,
                "Error al crear la categoria. Por favor, revisa los datos ingresados.",
            )
    return render(request, "formscategorias.html", {"form": form, "accion": "Crear"})

@login_required
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categorias, pk=pk)
    if request.method == "POST":
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria actualizada correctamente.")
            return redirect("lista_categorias")
        else:
            messages.error(
                request,
                "Error al actualizar la categoría. Por favor, revisa los datos ingresados.",
            )
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, "formscategorias.html", {"form": form, "accion": "Editar"})

@login_required
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categorias, pk=pk)
    categoria.delete()
    messages.success(request, "Categoria eliminada.")
    return redirect("lista_categorias")


@login_required
def estadisticas_publicaciones_por_categoria(request):
    publicaciones_por_categoria_query = (
        Publicaciones.objects.values("categoria__nombre")
        .annotate(total=Count("id"))
        .order_by("categoria__nombre")
    )

    categoria_labels = []
    categoria_data = []
    categoria_colors = []
    categoria_detalles = []

    colores_base_categorias = [
        "rgba(255, 99, 132, 0.7)",
        "rgba(54, 162, 235, 0.7)",
        "rgba(255, 206, 86, 0.7)",
        "rgba(75, 192, 192, 0.7)",
        "rgba(153, 102, 255, 0.7)",
        "rgba(255, 159, 64, 0.7)",
        "rgba(199, 199, 199, 0.7)",
        "rgba(83, 102, 255, 0.7)",
        "rgba(255, 59, 128, 0.7)",
        "rgba(100, 200, 100, 0.7)",
    ]

    for i, item in enumerate(publicaciones_por_categoria_query):
        label = item["categoria__nombre"]
        count = item["total"]
        categoria_labels.append(label)
        categoria_data.append(count)
        categoria_colors.append(
            colores_base_categorias[i % len(colores_base_categorias)]
        )
        categoria_detalles.append({"label": label, "count": count})

    # Para segundo grafico
    current_year = timezone.now().year
    publicaciones_por_mes_query = (
        Publicaciones.objects.filter(fecha_creacion__year=current_year)
        .annotate(month=ExtractMonth("fecha_creacion"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    mes_labels = []
    mes_data = []
    mes_colors = []
    mes_detalles = []

    nombres_meses = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }

    monthly_data_map = {
        item["month"]: item["count"] for item in publicaciones_por_mes_query
    }

    for i in range(1, 13):
        label = nombres_meses.get(i, f"Mes {i}")
        count = monthly_data_map.get(i, 0)
        mes_labels.append(label)
        mes_data.append(count)

        r = random.randint(0, 255)
        g = random.randint(0, 255)
        b = random.randint(0, 255)
        mes_colors.append(f"rgba({r}, {g}, {b}, 0.7)")
        mes_detalles.append({"label": label, "count": count})

    context = {
        "categoria_titulo_grafico": "Publicaciones por Categoría",
        "categoria_labels": categoria_labels,
        "categoria_data": categoria_data,
        "categoria_colors": categoria_colors,
        "categoria_chart_id": "publicacionesPorCategoriaChart",
        "categoria_chart_type": "bar",
        "categoria_detalles": categoria_detalles,
        "mes_titulo_grafico": f"Publicaciones por Mes ({current_year})",
        "mes_labels": mes_labels,
        "mes_data": mes_data,
        "mes_colors": mes_colors,
        "mes_chart_id": "publicacionesPorMesChart",
        "mes_chart_type": "pie",
        "mes_detalles": mes_detalles,
    }

    return render(request, "estadisticas_publicaciones_por_categoria.html", context)



@login_required
def exportar_estadisticas_excel(request):
    
    workbook = Workbook()
    
    sheet_categoria = workbook.active
    sheet_categoria.title = "Publicaciones por Categoría"

    headers_categoria = ["Categoría", "Total de Publicaciones"]
    sheet_categoria.append(headers_categoria)

    header_font = Font(bold=True)
    for cell in sheet_categoria[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    sheet_categoria.column_dimensions["A"].width = 30
    sheet_categoria.column_dimensions["B"].width = 25

    publicaciones_por_categoria_query = (
        Publicaciones.objects.values("categoria__nombre")
        .annotate(total=Count("id"))
        .order_by("categoria__nombre")
    )

    for item in publicaciones_por_categoria_query:
        sheet_categoria.append([item["categoria__nombre"], item["total"]])
        for cell in sheet_categoria[sheet_categoria.max_row]:
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    sheet_mes = workbook.create_sheet("Publicaciones por Mes")

    headers_mes = ["Mes", "Total de Publicaciones"]
    sheet_mes.append(headers_mes)

    for cell in sheet_mes[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    sheet_mes.column_dimensions["A"].width = 20
    sheet_mes.column_dimensions["B"].width = 25

    current_year = timezone.now().year
    publicaciones_por_mes_query = (
        Publicaciones.objects.filter(fecha_creacion__year=current_year)
        .annotate(month=ExtractMonth("fecha_creacion"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    nombres_meses = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }
    monthly_data_map = {
        item["month"]: item["count"] for item in publicaciones_por_mes_query
    }

    for i in range(1, 13):
        mes_nombre = nombres_meses.get(i, f"Mes {i}")
        count = monthly_data_map.get(i, 0)
        sheet_mes.append([mes_nombre, count])
        for cell in sheet_mes[sheet_mes.max_row]:
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=estadisticas_publicaciones.xlsx"
    )

    workbook.save(response)
    return response
